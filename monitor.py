import openpyxl
import requests
import os
import json
import re
import hashlib
from datetime import datetime, timedelta

# --- KONFIGURACJA ---
URL_STRONY = "https://uczelniaoswiecim.edu.pl/instytuty/new-instytut-nauk-inzynieryjno-technicznych/harmonogramy/"
TOKEN = os.getenv("TELEGRAM_TOKEN")
CHAT_ID = os.getenv("CHAT_ID")
STATE_FILE = "plan_state.json"
ICS_FILE = "studia.ics"
GRUPA = "MECH II"

PL_MONTHS = {"stycznia": 1, "lutego": 2, "marca": 3, "kwietnia": 4, "maja": 5, "czerwca": 6, "lipca": 7, "sierpnia": 8, "września": 9, "października": 10, "listopada": 11, "grudnia": 12}

def send_msg(text):
    if not TOKEN or not CHAT_ID: return
    url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
    requests.post(url, data={"chat_id": CHAT_ID, "text": text, "parse_mode": "Markdown"})

def get_merged_value(sheet, row, col):
    """Pobiera wartość nawet jeśli komórka jest scalona"""
    cell = sheet.cell(row=row, column=col)
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return cell.value

def parse_date(date_str):
    if not date_str: return None
    match = re.search(r"(\d+)\s+([a-zA-Z]+)", str(date_str))
    if match:
        day = int(match.group(1))
        month = PL_MONTHS.get(match.group(2).lower())
        if month: return datetime(2026, month, day)
    return None

def is_red(cell):
    if not cell.font or not cell.font.color: return False
    c = cell.font.color
    return str(c.rgb) in ["FFFF0000", "FF0000"] or c.indexed == 10

def main():
    print("🌐 Łączenie ze stroną...")
    r = requests.get(URL_STRONY)
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(r.text, 'html.parser')
    plan_url = next((a['href'] for a in soup.find_all('a', href=True) if "niestacjonarne" in a.text.lower() and a['href'].endswith('.xlsx')), None)
    
    if not plan_url:
        send_msg("❌ Błąd: Nie znaleziono linku do pliku Excel na stronie!")
        return

    resp = requests.get(plan_url)
    with open("temp.xlsx", "wb") as f: f.write(resp.content)
    wb = openpyxl.load_workbook("temp.xlsx", data_only=True)
    ws = wb.active

    # 1. Szukanie dat i kolumn grupy
    all_blocks = []
    mech_cols_global = []

    # Skanujemy wiersze nagłówkowe (1-5)
    for c in range(1, ws.max_column + 1):
        # Szukanie dat
        for r in range(1, 4):
            val = get_merged_value(ws, r, c)
            d = parse_date(val)
            if d and not any(b['date'] == d for b in all_blocks):
                all_blocks.append({"date": d, "col": c})
        
        # Szukanie MECH II
        for r in range(1, 6):
            val = str(get_merged_value(ws, r, c)).upper()
            if GRUPA in val:
                mech_cols_global.append(c)
                break

    if not all_blocks or not mech_cols_global:
        send_msg(f"❌ Diagnoza: Znaleziono dni: {len(all_blocks)}, kolumn MECH II: {len(mech_cols_global)}. Coś jest nie tak ze strukturą pliku.")
        return

    all_slots = []
    for i, block in enumerate(all_blocks):
        c_start = block["col"]
        c_end = all_blocks[i+1]["col"] if i+1 < len(all_blocks) else ws.max_column + 1
        target_cols = [c for c in mech_cols_global if c_start <= c < c_end]
        
        # Szukanie kolumny godzin (w obrębie bloku daty)
        h_col = None
        for test_c in range(c_start, c_start + 5):
            val = str(ws.cell(row=6, column=test_c).value).replace('.0','')
            if val.isdigit() and 7 <= int(val) <= 20:
                h_col = test_c
                break
        
        if not h_col: continue

        for row in range(5, ws.max_row + 1):
            try:
                # Godzina i Minuta
                h_val = ws.cell(row=row, column=h_col).value
                m_val = ws.cell(row=row, column=h_col+1).value
                if h_val is None or m_val is None: continue
                
                h = int(float(str(h_val)))
                m = int(float(str(m_val)))
                time_dt = block["date"].replace(hour=h, minute=m)
                
                for t_c in target_cols:
                    cell = ws.cell(row=row, column=t_c)
                    val = str(cell.value).strip()
                    if not val or val.lower() == "nan" or GRUPA in val.upper(): continue
                    
                    # Czyścimy tekst i sprawdzamy kolor
                    clean_val = " ".join(val.split())
                    if is_red(cell): clean_val = f"🔴 [ZDALNIE] {clean_val}"
                    all_slots.append({'start': time_dt, 'title': clean_val, 'col': t_c})
            except: continue

    # 2. Łączenie slotów
    all_slots.sort(key=lambda x: (x['col'], x['start']))
    merged_events = []
    if all_slots:
        curr = all_slots[0].copy()
        curr['end'] = curr['start'] + timedelta(minutes=45)
        for nxt in all_slots[1:]:
            if nxt['title'] == curr['title'] and nxt['col'] == curr['col'] and nxt['start'] == curr['end']:
                curr['end'] = nxt['start'] + timedelta(minutes=45)
            else:
                merged_events.append(curr)
                curr = nxt.copy()
                curr['end'] = curr['start'] + timedelta(minutes=45)
        merged_events.append(curr)

    if not merged_events:
        send_msg("⚠️ Skrypt przeanalizował plik, ale nie znalazł żadnych zajęć przypisanych do MECH II.")
        return

    # 3. Generowanie ICS
    ics = ["BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//PlanBot//MECHII", "X-WR-CALNAME:Plan MECH II", "METHOD:PUBLISH"]
    for e in merged_events:
        uid = hashlib.md5(f"{e['title']}{e['start']}{e['col']}".encode()).hexdigest()
        ics.extend(["BEGIN:VEVENT", f"UID:{uid}@studia.pl", f"DTSTAMP:{datetime.now().strftime('%Y%m%dT%H%M%SZ')}",
                    f"DTSTART:{e['start'].strftime('%Y%m%dT%H%M%S')}", f"DTEND:{e['end'].strftime('%Y%m%dT%H%M%S')}",
                    f"SUMMARY:{e['title']}", "END:VEVENT"])
    ics.append("END:VCALENDAR")
    
    with open(ICS_FILE, "w", encoding="utf-8") as f:
        f.write("\r\n".join(ics))

    # 4. Telegram Raport
    merged_events.sort(key=lambda x: x['start'])
    report = "🚨 *DOKŁADNY PLAN MECH II*\n"
    last_d = ""
    for e in merged_events:
        d_str = e['start'].strftime("%d.%m")
        if d_str != last_d:
            report += f"\n📅 *{d_str}*\n"
            last_d = d_str
        report += f"  • {e['start'].strftime('%H:%M')}-{e['end'].strftime('%H:%M')}: {e['title']}\n"
    
    send_msg(report)

if __name__ == "__main__":
    main()
